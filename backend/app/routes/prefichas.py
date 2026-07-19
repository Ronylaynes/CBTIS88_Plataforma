# backend/app/routes/prefichas.py
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.middleware.role_required import role_required
from app.middleware.validators import validate_preficha
import io
import base64

prefichas_bp = Blueprint('prefichas', __name__)

ROLES_REPORTE = ['admin', 'servicios_escolares']

# ════════════════════════════════════════════════════════════
#  CRUD BÁSICO
# ════════════════════════════════════════════════════════════

@prefichas_bp.route('', methods=['GET'])
@jwt_required()
def get_all_prefichas():
    from app.controllers.preficha_controller import PrefichaController
    page     = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    filters  = {
        'status':         request.args.get('status'),
        'payment_status': request.args.get('payment_status'),
        'search':         request.args.get('search'),
    }
    filters = {k: v for k, v in filters.items() if v is not None}
    result, status_code = PrefichaController.get_all_prefichas(
        page, per_page, filters)
    return jsonify(result), status_code

@prefichas_bp.route('', methods=['POST'])
def create_preficha():
    from app.controllers.preficha_controller import PrefichaController
    data = request.get_json()
    is_valid, errors = validate_preficha(data)
    if not is_valid:
        return jsonify({'errors': errors}), 400
    result, status_code = PrefichaController.create_preficha(data)
    return jsonify(result), status_code

@prefichas_bp.route('/<int:preficha_id>', methods=['GET'])
def get_preficha_by_id(preficha_id):
    from app.controllers.preficha_controller import PrefichaController
    result, status_code = PrefichaController.get_preficha_by_id(preficha_id)
    return jsonify(result), status_code

@prefichas_bp.route('/folio/<string:folio>', methods=['GET'])
def get_preficha_by_folio(folio):
    from app.controllers.preficha_controller import PrefichaController
    result, status_code = PrefichaController.get_preficha_by_folio(folio)
    return jsonify(result), status_code

@prefichas_bp.route('/<int:preficha_id>', methods=['PUT'])
@jwt_required()
def update_preficha(preficha_id):
    from app.controllers.preficha_controller import PrefichaController
    data = request.get_json()
    result, status_code = PrefichaController.update_preficha(preficha_id, data)
    return jsonify(result), status_code

@prefichas_bp.route('/<int:preficha_id>', methods=['DELETE'])
@jwt_required()
def delete_preficha(preficha_id):
    from app.controllers.preficha_controller import PrefichaController
    result, status_code = PrefichaController.delete_preficha(preficha_id)
    return jsonify(result), status_code

@prefichas_bp.route('/<int:preficha_id>/payment', methods=['PATCH'])
@jwt_required()
def update_payment_status(preficha_id):
    from app.controllers.preficha_controller import PrefichaController
    data = request.get_json()
    result, status_code = PrefichaController.update_payment_status(
        preficha_id, data)
    return jsonify(result), status_code

@prefichas_bp.route('/<int:preficha_id>/status', methods=['PATCH'])
@jwt_required()
@role_required(ROLES_REPORTE)
def update_status(preficha_id):
    from app.controllers.preficha_controller import PrefichaController
    data = request.get_json()
    result, status_code = PrefichaController.update_status(
        preficha_id, data)
    return jsonify(result), status_code

@prefichas_bp.route('/statistics', methods=['GET'])
@jwt_required()
def get_statistics():
    from app.controllers.preficha_controller import PrefichaController
    result, status_code = PrefichaController.get_statistics()
    return jsonify(result), status_code

# ════════════════════════════════════════════════════════════
#  PDF INDIVIDUAL
# ════════════════════════════════════════════════════════════

@prefichas_bp.route('/pdf/<string:folio>', methods=['GET'])
def descargar_pdf_preficha(folio):
    from app.models.preficha import Preficha
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import (SimpleDocTemplate, Table,
                                    TableStyle, Paragraph,
                                    Spacer, Image)
    from reportlab.lib        import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units  import cm

    preficha = Preficha.query.filter_by(folio=folio).first()
    if not preficha:
        return jsonify({'error': 'Preficha no encontrada'}), 404

    archivo = io.BytesIO()
    doc = SimpleDocTemplate(
        archivo, pagesize=letter,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=2*cm,  rightMargin=2*cm)
    elementos = []

    inst_style = ParagraphStyle(
        'inst', fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#666666'),
        alignment=1, spaceAfter=2)
    nombre_style = ParagraphStyle(
        'nombre', fontSize=11, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4A1220'),
        alignment=1, spaceAfter=2)
    sub_style = ParagraphStyle(
        'sub', fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#888888'),
        alignment=1, spaceAfter=2)
    titulo_style = ParagraphStyle(
        'titulo', fontSize=14, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4A1220'),
        alignment=1, spaceAfter=4)
    folio_style = ParagraphStyle(
        'folio', fontSize=9, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#C9963A'),
        alignment=1, spaceAfter=6)

    elementos.append(Paragraph(
        'Secretaría de Educación Pública · DGETI', inst_style))
    elementos.append(Paragraph(
        'Centro de Bachillerato Tecnológico Industrial '
        'y de Servicios No. 88', nombre_style))
    # ✅ Vicente Guerrero
    elementos.append(Paragraph('Vicente Guerrero', nombre_style))
    elementos.append(Paragraph('Tapachula, Chiapas', sub_style))
    elementos.append(Spacer(1, 0.3*cm))

    elementos.append(Table([['']], colWidths=[17*cm],
        style=TableStyle([
            ('LINEABOVE', (0,0), (-1,0), 1.5,
             colors.HexColor('#C9963A')),
            ('LINEBELOW', (0,0), (-1,0), 0.5,
             colors.HexColor('#C9963A')),
        ])))
    elementos.append(Spacer(1, 0.3*cm))

    # ✅ Título actualizado
    elementos.append(Paragraph(
        'FICHA DE APLICACIÓN DE INSTRUMENTO DIAGNÓSTICO',
        titulo_style))
    elementos.append(Paragraph(
        f'Folio: {preficha.folio}  ·  '
        f'Fecha: {str(preficha.created_at)[:10] if preficha.created_at else "—"}',
        folio_style))
    elementos.append(Spacer(1, 0.3*cm))

    if preficha.foto_base64 and ',' in preficha.foto_base64:
        try:
            _, b64data  = preficha.foto_base64.split(',', 1)
            img_bytes   = base64.b64decode(b64data)
            img_stream  = io.BytesIO(img_bytes)
            foto        = Image(img_stream, width=3*cm, height=3.5*cm)
            foto.hAlign = 'CENTER'
            elementos.append(foto)
            elementos.append(Spacer(1, 0.3*cm))
        except Exception:
            pass

    VINO   = colors.HexColor('#4A1220')
    DORADO = colors.HexColor('#C9963A')
    CREMA  = colors.HexColor('#FDF8F0')
    CREMA2 = colors.HexColor('#F5E6C8')

    datos = [
        ['I. DATOS PERSONALES', ''],
        ['Nombre completo',
         f"{preficha.nombre or ''} "
         f"{preficha.apellido_paterno or ''} "
         f"{preficha.apellido_materno or ''}"],
        ['CURP',              preficha.curp or '—'],
        ['Fecha nacimiento',
         str(preficha.fecha_nacimiento)
         if preficha.fecha_nacimiento else '—'],
        ['Sexo',
         preficha.sexo.value if preficha.sexo else '—'],
        ['II. CONTACTO', ''],
        ['Correo',            preficha.email    or '—'],
        ['Teléfono',          preficha.telefono or '—'],
        ['Domicilio',         preficha.direccion or '—'],
        ['Nombre del tutor',  preficha.tutor_nombre or '—'],
        ['Parentesco',        preficha.tutor_parentesco or '—'],
        ['Tel. tutor',        preficha.tutor_telefono or '—'],
        ['III. OPCIONES ACADÉMICAS', ''],
        ['1ª Especialidad',   preficha.especialidad_1 or '—'],
        ['2ª Especialidad',   preficha.especialidad_2 or '—'],
        ['3ª Especialidad',   preficha.especialidad_3 or '—'],
        ['Secundaria',        preficha.secundaria_nombre or '—'],
        ['Promedio',
         str(preficha.promedio_egreso)
         if preficha.promedio_egreso else '—'],
        ['IV. ESTATUS', ''],
        ['Status',
         preficha.status.value
         if preficha.status else 'pendiente'],
        ['Observaciones',     preficha.observaciones or '—'],
    ]

    secciones_idx = [0, 5, 12, 18]
    tabla = Table(datos, colWidths=[6*cm, 11*cm])
    style_cmds = [
        ('FONTSIZE',    (0,0), (-1,-1), 9),
        ('PADDING',     (0,0), (-1,-1), 5),
        ('GRID',        (0,0), (-1,-1), 0.4,
         colors.HexColor('#D4B89A')),
        ('FONTNAME',    (0,0), (0,-1),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [CREMA, CREMA2]),
    ]
    for idx in secciones_idx:
        style_cmds += [
            ('BACKGROUND', (0,idx), (-1,idx), VINO),
            ('TEXTCOLOR',  (0,idx), (-1,idx), DORADO),
            ('FONTNAME',   (0,idx), (-1,idx), 'Helvetica-Bold'),
            ('SPAN',       (0,idx), (1, idx)),
            ('ALIGN',      (0,idx), (-1,idx), 'CENTER'),
        ]
    tabla.setStyle(TableStyle(style_cmds))
    elementos.append(tabla)
    elementos.append(Spacer(1, 0.5*cm))

    firma_style = ParagraphStyle(
        'firma', fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#888888'), alignment=1)
    elementos.append(Table([['']], colWidths=[17*cm],
        style=TableStyle([
            ('LINEABOVE', (0,0), (-1,0), 0.5,
             colors.HexColor('#C9963A')),
        ])))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(
        'CBTIS No. 88  ·  Vicente Guerrero  ·  '
        'Tapachula, Chiapas  ·  Ciclo Escolar 2025–2026',
        firma_style))
    elementos.append(Paragraph(
        'Documento generado automáticamente por el '
        'Sistema de Prefichas del CBTIS No. 88.',
        firma_style))

    doc.build(elementos)
    archivo.seek(0)
    return send_file(
        archivo, mimetype='application/pdf',
        as_attachment=True,
        download_name=f'ficha_diagnostico_{folio}.pdf')

# ════════════════════════════════════════════════════════════
#  REPORTE EXCEL — Solo admin y servicios_escolares
# ════════════════════════════════════════════════════════════

@prefichas_bp.route('/reporte/excel', methods=['GET'])
@jwt_required()
@role_required(ROLES_REPORTE)
def descargar_excel_prefichas():
    from app.models.preficha import Preficha
    import openpyxl
    from openpyxl.styles import (Font, PatternFill,
                                  Alignment, Border, Side)
    from openpyxl.utils  import get_column_letter
    from datetime        import datetime

    prefichas = Preficha.query.order_by(
        Preficha.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relación de Aspirantes'

    VINO   = '6B1E2E'
    DORADO = 'C9963A'
    BLANCO = 'FFFFFF'
    GRIS   = 'F5F5F5'

    borde = Border(
        left   = Side(style='thin', color='CCCCCC'),
        right  = Side(style='thin', color='CCCCCC'),
        top    = Side(style='thin', color='CCCCCC'),
        bottom = Side(style='thin', color='CCCCCC'),
    )

    # Fila 1 — Encabezado institucional
    ws.merge_cells('A1:R1')
    ws['A1'] = 'CBTIS No. 88 — Vicente Guerrero — Tapachula, Chiapas'
    ws['A1'].font      = Font(bold=True, size=14,
                              color=BLANCO, name='Calibri')
    ws['A1'].fill      = PatternFill('solid', fgColor=VINO)
    ws['A1'].alignment = Alignment(horizontal='center',
                                   vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2 — Subtítulo
    ws.merge_cells('A2:R2')
    ws['A2'] = (
        'RELACIÓN DE ASPIRANTES — '
        'FICHA DE APLICACIÓN DE INSTRUMENTO DIAGNÓSTICO — '
        f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    ws['A2'].font      = Font(size=9, color=BLANCO, name='Calibri')
    ws['A2'].fill      = PatternFill('solid', fgColor='4A1220')
    ws['A2'].alignment = Alignment(horizontal='center',
                                   vertical='center')
    ws.row_dimensions[2].height = 16

    # Fila 3 — Total
    ws.merge_cells('A3:R3')
    ws['A3'] = f'Total de aspirantes registrados: {len(prefichas)}'
    ws['A3'].font      = Font(bold=True, size=10,
                              color=VINO, name='Calibri')
    ws['A3'].fill      = PatternFill('solid', fgColor='FFF3E0')
    ws['A3'].alignment = Alignment(horizontal='left',
                                   vertical='center')
    ws.row_dimensions[3].height = 16

    # Fila 4 — Encabezados columnas
    # ✅ Sin Método Pago ni Pago Confirmado
    columnas = [
        ('Folio',            14),
        ('CURP',             22),
        ('Nombre',           18),
        ('Apellido Paterno', 18),
        ('Apellido Materno', 18),
        ('Nombre Completo',  28),
        ('Fecha Nacimiento', 16),
        ('Sexo',              8),
        ('Teléfono',         14),
        ('Correo',           28),
        ('Dirección',        30),
        ('1ª Especialidad',  22),
        ('2ª Especialidad',  22),
        ('3ª Especialidad',  22),
        ('Secundaria',       24),
        ('Promedio',         10),
        ('Status',           14),
        ('Fecha Registro',   18),
    ]

    for col_idx, (nombre, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=4, column=col_idx, value=nombre)
        celda.font      = Font(bold=True, size=10,
                               color=BLANCO, name='Calibri')
        celda.fill      = PatternFill('solid', fgColor=DORADO)
        celda.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    wrap_text=True)
        celda.border    = borde
        ws.column_dimensions[
            get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[4].height = 22

    # Filas de datos
    for fila_idx, p in enumerate(prefichas, start=5):
        color_fila = GRIS if fila_idx % 2 == 0 else BLANCO
        fill_fila  = PatternFill('solid', fgColor=color_fila)

        # ✅ Sin metodo_pago ni pago_confirmado
        fila_data = [
            p.folio             or '—',
            p.curp              or '—',
            p.nombre            or '—',
            p.apellido_paterno  or '—',
            p.apellido_materno  or '—',
            p.nombre_completo   or '—',
            str(p.fecha_nacimiento)
            if p.fecha_nacimiento else '—',
            p.sexo.value        if p.sexo   else '—',
            p.telefono          or '—',
            p.email             or '—',
            p.direccion         or '—',
            p.especialidad_1    or '—',
            p.especialidad_2    or '—',
            p.especialidad_3    or '—',
            p.secundaria_nombre or '—',
            float(p.promedio_egreso)
            if p.promedio_egreso else '—',
            p.status.value      if p.status else 'pendiente',
            str(p.created_at)[:10]
            if p.created_at else '—',
        ]

        for col_idx, valor in enumerate(fila_data, start=1):
            celda = ws.cell(row=fila_idx,
                            column=col_idx, value=valor)
            celda.font      = Font(size=9, name='Calibri')
            celda.fill      = fill_fila
            celda.alignment = Alignment(horizontal='left',
                                        vertical='center')
            celda.border    = borde
        ws.row_dimensions[fila_idx].height = 15

    ws.freeze_panes = 'A5'

    archivo = io.BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    fecha_hoy = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        archivo,
        mimetype=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'),
        as_attachment=True,
        download_name=f'relacion_aspirantes_{fecha_hoy}.xlsx')

# ════════════════════════════════════════════════════════════
#  REPORTE CSV — Solo admin y servicios_escolares
# ════════════════════════════════════════════════════════════

@prefichas_bp.route('/reporte/csv', methods=['GET'])
@jwt_required()
@role_required(ROLES_REPORTE)
def descargar_csv_prefichas():
    from app.models.preficha import Preficha
    import csv
    from datetime import datetime

    prefichas = Preficha.query.order_by(
        Preficha.created_at.desc()).all()

    archivo = io.StringIO()
    archivo.write('\ufeff')  # BOM para Excel de Windows
    writer  = csv.writer(archivo)

    # ✅ Sin Método Pago ni Pago Confirmado
    writer.writerow([
        'Folio', 'CURP', 'Nombre', 'Apellido Paterno',
        'Apellido Materno', 'Nombre Completo',
        'Fecha Nacimiento', 'Sexo', 'Teléfono', 'Correo',
        'Dirección', '1ª Especialidad', '2ª Especialidad',
        '3ª Especialidad', 'Secundaria', 'Promedio',
        'Status', 'Fecha Registro',
    ])

    for p in prefichas:
        writer.writerow([
            p.folio             or '',
            p.curp              or '',
            p.nombre            or '',
            p.apellido_paterno  or '',
            p.apellido_materno  or '',
            p.nombre_completo   or '',
            str(p.fecha_nacimiento)
            if p.fecha_nacimiento else '',
            p.sexo.value        if p.sexo   else '',
            p.telefono          or '',
            p.email             or '',
            p.direccion         or '',
            p.especialidad_1    or '',
            p.especialidad_2    or '',
            p.especialidad_3    or '',
            p.secundaria_nombre or '',
            float(p.promedio_egreso)
            if p.promedio_egreso else '',
            p.status.value      if p.status else 'pendiente',
            str(p.created_at)[:10]
            if p.created_at else '',
        ])

    fecha_hoy = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(
        io.BytesIO(archivo.getvalue().encode('utf-8-sig')),
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name=f'relacion_aspirantes_{fecha_hoy}.csv')